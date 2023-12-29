from openpyxl import load_workbook
from HttpUtils import post
from datetime import datetime
import Config as config


def price_guangdong():
    current_datetime = datetime.now()
    to_day = current_datetime.strftime("%Y-%m-%d")
    excel_column_name = current_datetime.strftime("%y年%m月%d日")
    url = config.demo_url
    params = {
        "dayReal": "01",
        "operateDate": to_day,
        "tradeType": "03",
        "tranCode": "A012",
        "transCountInfoId": config.demo_transCountInfoId,
    }

    header = {"Cookie": config.cookie}
    json_data = post(url, params, header)
    print(json_data)

    # excel列与数据的映射关系
    position_map = {}
    for i in range(2, 26):
        value = '{:02d}:00'.format((i - 2) % 24)
        position_map[i] = value + "@" + "priceResultTwenty"
        position_map[i + 26] = value + "@" + "priceRecTwenty"

    # 读取 Excel 文件
    excel_path = config.demo_excel_path
    wb = load_workbook(excel_path)
    ws = wb[config.demo_sheet_name]

    # 获取所有列名
    #columns_name = [col[0].value for col in ws.iter_cols(min_row=1, max_col=ws.max_column)]

    data_map = {item["time"]: item for item in json_data["data"]}

    # 行标
    row_number = ws.max_row + 1
    # 先写第一列
    ws.cell(row=row_number, column=1, value=excel_column_name)
    # 遍历map，插入新数据
    for key, value in position_map.items():
        data_index = value.split("@")
        item = data_map.get(data_index[0])
        ws.cell(row=row_number, column=key, value=float(item[data_index[1]]))

    # 保存
    wb.save(excel_path)
