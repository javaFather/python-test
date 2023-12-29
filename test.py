from openpyxl import load_workbook
from datetime import datetime

# Create an instance of the Httputils class
# JSON 数据
json_data = {"code": 0, "msg": "success", "data": [
    {"priceResultTwenty": "353.46", "priceSendTwenty": "332.8325", "isPriorityPlanRec": "1",
     "powerRecTwenty": "594.748", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "00:00",
     "powerSendTwenty": "636.093", "priceRecTwenty": "446.89"},
    {"priceResultTwenty": "338.6", "priceSendTwenty": "335.5525", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1036.533", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "01:00",
     "powerSendTwenty": "1108.591", "priceRecTwenty": "431"},
    {"priceResultTwenty": "333.115", "priceSendTwenty": "316.9775", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1418.343", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "02:00",
     "powerSendTwenty": "1516.943", "priceRecTwenty": "425.135"},
    {"priceResultTwenty": "318.365", "priceSendTwenty": "329.4675", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1339.978", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "03:00",
     "powerSendTwenty": "1433.128", "priceRecTwenty": "409.355"},
    {"priceResultTwenty": "306.9", "priceSendTwenty": "322.4575", "isPriorityPlanRec": "1", "powerRecTwenty": "20.365",
     "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "04:00", "powerSendTwenty": "21.781",
     "priceRecTwenty": "397.09"},
    {"priceResultTwenty": "311.8975", "priceSendTwenty": "321.6475", "isPriorityPlanRec": "1",
     "powerRecTwenty": "820.978", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "05:00",
     "powerSendTwenty": "878.053", "priceRecTwenty": "402.4375"},
    {"priceResultTwenty": "357.3225", "priceSendTwenty": "342.15", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1940.863", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "06:00",
     "powerSendTwenty": "2075.79", "priceRecTwenty": "451.0225"},
    {"priceResultTwenty": "402.4625", "priceSendTwenty": "367.1775", "isPriorityPlanRec": "1",
     "powerRecTwenty": "2719.118", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "07:00",
     "powerSendTwenty": "2908.148", "priceRecTwenty": "499.3025"},
    {"priceResultTwenty": "506.2075", "priceSendTwenty": "409.935", "isPriorityPlanRec": "1",
     "powerRecTwenty": "3144.134", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "08:00",
     "powerSendTwenty": "3362.708", "priceRecTwenty": "610.2575"},
    {"priceResultTwenty": "548.2075", "priceSendTwenty": "414.185", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1608.834", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "09:00",
     "powerSendTwenty": "1720.679", "priceRecTwenty": "655.1775"},
    {"priceResultTwenty": "577.5525", "priceSendTwenty": "412.4325", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1055.588", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "10:00",
     "powerSendTwenty": "1128.973", "priceRecTwenty": "686.5625"},
    {"priceResultTwenty": "547.435", "priceSendTwenty": "405.545", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1680.71", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "11:00",
     "powerSendTwenty": "1797.553", "priceRecTwenty": "654.355"},
    {"priceResultTwenty": "364.8675", "priceSendTwenty": "341.51", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1130.148", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "12:00",
     "powerSendTwenty": "1208.717", "priceRecTwenty": "459.0875"},
    {"priceResultTwenty": "408.015", "priceSendTwenty": "364.66", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1929.146", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "13:00",
     "powerSendTwenty": "2063.255", "priceRecTwenty": "505.245"},
    {"priceResultTwenty": "519.21", "priceSendTwenty": "389.4625", "isPriorityPlanRec": "1",
     "powerRecTwenty": "1703.581", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "14:00",
     "powerSendTwenty": "1822.011", "priceRecTwenty": "624.17"},
    {"priceResultTwenty": "525.2775", "priceSendTwenty": "407.9225", "isPriorityPlanRec": "1",
     "powerRecTwenty": "3268.594", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "15:00",
     "powerSendTwenty": "3495.819", "priceRecTwenty": "630.6575"},
    {"priceResultTwenty": "579.1625", "priceSendTwenty": "413.875", "isPriorityPlanRec": "1",
     "powerRecTwenty": "2625.346", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "16:00",
     "powerSendTwenty": "2807.859", "priceRecTwenty": "688.2825"},
    {"priceResultTwenty": "599.6825", "priceSendTwenty": "491.025", "isPriorityPlanRec": "1",
     "powerRecTwenty": "2191.678", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "17:00",
     "powerSendTwenty": "2344.04", "priceRecTwenty": "710.2325"},
    {"priceResultTwenty": "694.7025", "priceSendTwenty": "703.1825", "isPriorityPlanRec": "1",
     "powerRecTwenty": "2766.795", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "18:00",
     "powerSendTwenty": "2959.14", "priceRecTwenty": "811.8625"},
    {"priceResultTwenty": "704.9475", "priceSendTwenty": "638.47", "isPriorityPlanRec": "1",
     "powerRecTwenty": "3436.125", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "19:00",
     "powerSendTwenty": "3675", "priceRecTwenty": "822.8175"},
    {"priceResultTwenty": "717.4425", "priceSendTwenty": "624.49", "isPriorityPlanRec": "1", "powerRecTwenty": "4581.5",
     "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "20:00", "powerSendTwenty": "4900",
     "priceRecTwenty": "836.1825"},
    {"priceResultTwenty": "632.6025", "priceSendTwenty": "545.7425", "isPriorityPlanRec": "1",
     "powerRecTwenty": "4174.99", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "21:00",
     "powerSendTwenty": "4465.23", "priceRecTwenty": "745.4425"},
    {"priceResultTwenty": "536.7225", "priceSendTwenty": "420.0225", "isPriorityPlanRec": "1",
     "powerRecTwenty": "3460.396", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "22:00",
     "powerSendTwenty": "3700.956", "priceRecTwenty": "642.8925"},
    {"priceResultTwenty": "441.675", "priceSendTwenty": "390.265", "isPriorityPlanRec": "1",
     "powerRecTwenty": "2299.674", "isPriorityPlanSend": "1", "loseResultTwenty": 68.86, "time": "23:00",
     "powerSendTwenty": "2459.546", "priceRecTwenty": "541.235"}]}

# excel列与数据的映射关系
position_map = {}
for i in range(2, 26):
    value = '{:02d}:00'.format((i - 2) % 24)
    position_map[i] = value + "@" + "priceResultTwenty"
    position_map[i + 26] = value + "@" + "priceRecTwenty"

# for key, value in position_map.items():
#     print(f'Key: {key}, Value: {value}')

# 读取 Excel 文件
excel_path = "C:/Users/wangzhenxing-jk/Desktop/工作簿2.xlsx"
wb = load_workbook(excel_path)
ws = wb["日前溪右送出侧电量和价格广东"]

# 获取所有列名
columns_name = [col[0].value for col in ws.iter_cols(min_row=1, max_col=ws.max_column)]

data_map = {item["time"]: item for item in json_data["data"]}

# 行标
row_number = ws.max_row + 1
current_datetime = datetime.now()
to_day = current_datetime.strftime("%Y-%m-%d")
excel_column_name = current_datetime.strftime("%y年%m月%d日")
# 遍历map，插入新数据
ws.cell(row=row_number, column=1, value=excel_column_name)
for key, value in position_map.items():
    data_index = value.split("@")
    item = data_map.get(data_index[0])
    cell = ws.cell(row=row_number, column=key, value=float(item[data_index[1]]))

# 保存
wb.save(excel_path)
